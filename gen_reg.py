#########################################################################################
# Engineer:         czz
# Date:             2025-03-18
# Version:          1.0
# Description: This script is used to generate Verilog register files from an Excel file.
#########################################################################################

import openpyxl
import re
from datetime import datetime
import argparse
def extract_offset_components(expression):
    pattern = r'''
        ^\s*
        (0[Xx][0-9a-fA-F]+)        # Base address (group 1)
        (?:                      # Optional variable-step section
            \s*\+\s*
            ([a-zA-Z_]\w*)       # Variable name (group 2)
            \s*\*\s*
            (0[Xx][0-9a-fA-F]+)     # Step value (group 3)
        )?
        \s*$
    '''
    
    match = re.match(pattern, expression, re.VERBOSE)
    if not match:
        raise ValueError(f"Invalid expression: {expression}")

    base_addr = match.group(1)
    variable = match.group(2) if match.group(2) else None
    step_value = match.group(3) if match.group(3) else None
    
    return base_addr, variable, step_value

class Register:
    def __init__(self):
        self.offset = ""
        self.var = ""
        self.var_step = ""
        self.reg_name = ""
        self.var_val= ""
        self.fields = []
        self.wr = ""

class Field:
    def __init__(self):
        self.bits = ""
        self.bits_size = ""
        self.name = ""
        self.sw_access = ""
        self.default = ""

def parse_excel_var(sheet):
    result = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Stop at first empty row
        if not row[0]:  
            break
            
        name = row[0]
        range_str = row[1]
        ___, max_val = map(int, range_str.split('~'))
        result[name] = max_val
    return result

def calculate_bit_width(bit_str):
    bit_str = str(bit_str)  # fix: ensure bit_str is a string
    parts = bit_str.split(':')
    if len(parts) == 1:
        _ = int(parts[0])  # Check if it is a valid integer
        return 1
        
    if len(parts) == 2:
        start = int(parts[0])
        end = int(parts[1])
        return abs(start - end) + 1 
            
    raise ValueError("format bits error: %s" % bit_str)

def parse_excel(filename,parallel=False):
    wb = openpyxl.load_workbook(filename)
    module_info = {}
    registers = []
    
    ws = wb.worksheets[1]
    var_ranges = parse_excel_var(ws)

    # Parse module information
    ws = wb.worksheets[0]

    for row in ws.iter_rows(max_row=9, values_only=True):   # Only read the first 10 rows
        if row[0] == "module":
            module_info["module"] = row[1]
            if module_info["module"] not in filename:
                print(f"Warning: Module name '{module_info['module']}' does not match filename '{filename}'. Changing module name to match filename.")
                module_info["module"] = filename.split('.')[0]  # Use filename as module name
        elif row[0] in ["owner", "size", "base_addr", "addr_width", "data_width", "cfg_interface"]:
            module_info[row[0]] = row[1]
    
    # Parse registers
    current_reg = None
    check_file = False
    # Convert module_size to bytes if it is given as a string like "4KB"
    module_size_str = module_info.get("size", "0")
    if module_size_str.endswith("KB"):
        module_size = int(module_size_str[:-2]) * 1024
    elif module_size_str.endswith("B"):
        module_size = int(module_size_str[:-1])
    else:
        module_size = int(module_size_str, 16)  # Assume it is in hexadecimal if no unit is specified

    offset_set = set()
    reg_name_set = set()
    field_name_set = set()

    for row in ws.iter_rows(min_row=10, values_only=True):

        if check_file == False:
            if row[0] != "offset":
                raise ValueError("Invalid register description file format. 'offset' column not found.")
            else:
                check_file = True
                continue

        if row[0]:  # New register
            offset, var, var_step = extract_offset_components(row[0])
            if int(offset, 16) > module_size:
                print(f"Warning: Offset '{offset}' is greater than module size: {module_size}B. Skipping register.")
                continue  # Skip parsing if offset is greater than module size

            if offset in offset_set:
                raise ValueError(f"Duplicate offset found: {offset}")
            offset_set.add(offset)

            if current_reg:
                # Check if all fields are RO before adding to registers
                all_ro = all(field.sw_access == 'RO' for field in current_reg.fields)
                current_reg.wr = 'r' if all_ro else 'w'
                registers.append(current_reg)

            current_reg = Register()
            current_reg.offset = offset[2:]
            current_reg.var_step = var_step[2:] if var_step else None
            current_reg.var = var

            if current_reg.var is not None:
                if current_reg.var in var_ranges:
                    current_reg.var_val = var_ranges[current_reg.var]
                else:
                    raise ValueError(f"Variable '{current_reg.var}' not found in variable ranges")
            if row[1] is None:
                raise ValueError("Register name not found")
            if row[1] in reg_name_set:
                raise ValueError(f"Duplicate register name found: {row[1]}")
            reg_name_set.add(row[1])
            current_reg.reg_name = row[1]

        if row[2]:  # Field
            required_columns = [2, 3, 4, 6]
            if any(row[i] is None for i in required_columns):
                raise ValueError("One or more required fields are missing in the '{required_columns(i)}' row")
            field = Field()
            field.bits = row[2]
            field.bits_size = calculate_bit_width(row[2])
            field.name = row[3]
            if field.name in field_name_set:
                raise ValueError(f"Duplicate field name found: {field.name}")
            field_name_set.add(field.name)
            field.sw_access = row[4].upper()  # Convert to uppercase
            if field.sw_access not in ['RW', 'W1P', 'W1C', 'RO']:
                raise ValueError(f"Invalid sw_access value: {field.sw_access}")
            field.default = row[6]
            current_reg.fields.append(field)

    if current_reg:
        # Check if all fields are RO before adding to registers
        all_ro = all(field.sw_access == 'RO' for field in current_reg.fields)
        current_reg.wr = 'r' if all_ro else 'w'
        registers.append(current_reg)
    
    if parallel == True:
        expanded_registers = []
        for reg in registers:
            if reg.var is not None:
                for i in range(reg.var_val + 1):
                    new_reg = Register()
                    new_reg.offset = hex(int(reg.offset, 16) + i * int(reg.var_step, 16))[2:]
                    new_reg.var_step = None
                    new_reg.var = None
                    new_reg.var_val = None
                    new_reg.reg_name = f"{reg.var}{i}_{reg.reg_name}"
                    new_reg.wr = reg.wr
                    for field in reg.fields:
                        new_field = Field()
                        new_field.bits = field.bits
                        new_field.bits_size = field.bits_size
                        new_field.name = f"{reg.var}{i}_{field.name}"
                        new_field.sw_access = field.sw_access
                        new_field.default = field.default
                        new_reg.fields.append(new_field)
                    expanded_registers.append(new_reg)
            else:
                expanded_registers.append(reg)
        expanded_registers.sort(key=lambda reg: int(reg.offset, 16))
        registers = expanded_registers

    return module_info, registers

def generate_verilog(module_info, registers, filename):
    now = datetime.now()
    code = []
    
    _addr_width = int(module_info.get('addr_width', '12'))
    # Header
    code.append(f"// Filename          : {filename}.v")
    code.append(f"// Author            : {module_info.get('owner', 'unknown')}")
    code.append(f"// Created           : {now.strftime('%Y-%m-%d %H:%M:%S')}")
    code.append( "// Description       : This file is auto generated by gen_reg.py script. Not edit by hand")
    code.append(f"//                   : addr_width = {_addr_width}")
    code.append(f"//                   : bus_type   = {module_info.get('cfg_interface', 'regbus')}")
    code.append(f"//                   : base_addr  = {module_info.get('base_addr', '32\'h0000')}\n")
    
    # Module declaration
    code.append(f"module {module_info['module']} (")
    code.append( "input               clk,")
    code.append( "input               rst_n,")
    code.append(f"input       [{_addr_width-1}:0]  reg_addr,")
    code.append( "input               wr_en,")
    code.append( "input               rd_en,")
    code.append( "input       [3 :0]  wr_msk,")
    code.append( "input       [31:0]  wr_data,")
    code.append( "output reg  [31:0]  rd_data,")
    
    # Signals
    for reg in registers:
        if reg.var is not None:# Variable registers
            for field in reg.fields:
                if field.sw_access == "RW" or field.sw_access == "W1P": 
                    if field.bits_size == 1:
                        code.append(f"output reg          {field.name}[{reg.var_val}:0],")
                    else:
                        code.append(f"output reg  [{field.bits_size-1:02}:0]  {field.name}[{reg.var_val}:0],")
                elif field.sw_access == "W1C":
                    if field.bits_size == 1:
                        code.append(f"input               {field.name}_hw_en[{reg.var_val}:0],")
                        code.append(f"input               {field.name}_hw_val[{reg.var_val}:0],")
                        code.append(f"output reg          {field.name},")
                    else:
                        code.append(f"input               {field.name}_hw_en[{reg.var_val}:0],")
                        code.append(f"input       [{field.bits_size-1:02}:0]  {field.name}_hw_val[{reg.var_val}:0],")
                        code.append(f"output reg  [{field.bits_size-1:02}:0]  {field.name}[{reg.var_val}:0],")
                elif field.sw_access == "RO":
                    if field.bits_size == 1:
                        code.append(f"input               {field.name}[{reg.var_val}:0],")
                    else:
                        code.append(f"input       [{field.bits_size-1:02}:0]  {field.name}[{reg.var_val}:0],")
        else:# Non-variable registers                        
            for field in reg.fields:
                if field.sw_access == "RW" or field.sw_access == "W1P": 
                    if field.bits_size == 1:
                        code.append(f"output reg          {field.name},")
                    else:
                        code.append(f"output reg  [{field.bits_size-1:02d}:0]  {field.name},")
                elif field.sw_access == "W1C":
                    if field.bits_size == 1:
                        code.append(f"input               {field.name}_hw_en,")
                        code.append(f"input               {field.name}_hw_val,")
                        code.append(f"output reg          {field.name},")
                    else:
                        code.append(f"input               {field.name}_hw_en,")
                        code.append(f"input       [{field.bits_size-1:02}:0]  {field.name}_hw_val,")
                        code.append(f"output reg  [{field.bits_size-1:02}:0]  {field.name},")
                elif field.sw_access == "RO":
                    if field.bits_size == 1:
                        code.append(f"input               {field.name},")
                    else:
                        code.append(f"input       [{field.bits_size-1:02}:0]  {field.name},")
    
    code[-1] = code[-1].rstrip(',')  # Remove last comma
    code.append(");\n")
    
    # Register and wire declarations
    code.append("//============================================================================")
    code.append("// reg and wire declaration")
    code.append("//============================================================================")
    code.append("reg  [31:0]    rd_data_nxt ;")
    code.append("wire [31:0]    msk;")
    
    # Address decode wires
    for reg in registers:
        if reg.var is not None:    
            if reg.wr == 'r':
                code.append(f"reg  [31:0]    rd_data_nxt_{reg.reg_name};")
            else:
                code.append(f"reg  [31:0]    rd_data_nxt_{reg.reg_name};")
                code.append(f"wire           wr_en_{reg.reg_name}[{reg.var_val}:0];")
        else:
            if reg.wr == 'w':
                code.append(f"wire           wr_en_{reg.reg_name};")

    # Main code
    code.append("//============================================================================")
    code.append("//main code")
    code.append("//============================================================================")
    code.append("assign msk = {{8{wr_msk[3]}},{8{wr_msk[2]}},{8{wr_msk[1]}},{8{wr_msk[0]}}};\n")
    
    # Address decoding
    code.append("//============================================================================")
    code.append("// reg wr_en/rd_en assignment")
    code.append("//============================================================================")

    gvar_i_declared = False
    ginter_j_declared = False
    # Calculate the maximum length of the left-hand side expressions
    max_lhs_length = 0
    for reg in registers:
        if reg.wr == 'w':
            if reg.var is not None:
                lhs_length = len(f"wr_en_{reg.reg_name}[{reg.var_val}]")
            else:
                lhs_length = len(f"wr_en_{reg.reg_name}")
            if lhs_length > max_lhs_length:
                max_lhs_length = lhs_length


    # Generate the Verilog code with aligned assign statements
    for reg in registers:
        if reg.wr == 'w':
            if reg.var is not None:
                if not gvar_i_declared:
                    code.append("\ngenvar i;")
                    gvar_i_declared = True
                code.append("generate")
                code.append(f"    for(i = 0; i <= {reg.var_val}; i = i + 1) begin: wr_{reg.reg_name}")
                code.append(f"        assign wr_en_{reg.reg_name}[i]= wr_en & (reg_addr[{_addr_width-1}:0] == {_addr_width}'h{reg.offset} + {_addr_width}'h{reg.var_step} * i );")
                code.append(f"    end")
                code.append("endgenerate")
            else:
                lhs = f"wr_en_{reg.reg_name}"
                code.append(f"assign {lhs:<{max_lhs_length}} = wr_en & (reg_addr[{_addr_width-1}:0] == {_addr_width}'h{reg.offset});")
    


    # Register writes
    code.append("\n//============================================================================")
    code.append(  "// reg write")
    code.append(  "//============================================================================")
    for reg in registers:
        if reg.var is not None:# Variable registers
            if reg.wr == 'w':
                for field in reg.fields:
                    code.append( "//============================================================================")
                    code.append(f"// {field.name} addr:{_addr_width}'h{reg.offset} type:{field.sw_access} bits:[{field.bits}] default:{field.default}")
                    code.append("//============================================================================")
                    code.append("generate")
                    code.append(f"    for(i = 0; i <= {reg.var_val}; i = i + 1) begin: wr_{field.name}")
                    code.append(f"        always @(posedge clk or negedge rst_n) begin")
                    code.append(f"            if (!rst_n)")
                    code.append(f"                {field.name}[i] <= {field.default};")
                    code.append(f"            else begin")
                    if field.sw_access == "RW":
                        code.append(f"                if (wr_en_{reg.reg_name}[i] == 1'b1)")
                        if field.bits_size == 1:
                            code.append(f"                    {field.name}[i] <= ({field.name}[i] & ~msk[{field.bits}]) | (wr_data[{field.bits}] & msk[{field.bits}]);")
                        else:
                            code.append(f"                    {field.name}[i][{field.bits_size-1}:0] <= ({field.name}[i][{field.bits_size-1}:0] & ~msk[{field.bits}]) | (wr_data[{field.bits}] & msk[{field.bits}]);")
                    elif field.sw_access == "W1P":
                        code.append(f"                if (wr_en_{reg.reg_name}[i] == 1'b1)")
                        code.append(f"                    {field.name}[i] <= wr_data[{field.bits}] & msk[{field.bits}];")
                    elif field.sw_access == "W1C":
                        code.append(f"                if ({field.name}_hw_en == 1'b1)")
                        code.append(f"                    {field.name}[i] <= {field.name}_hw_val;")
                        code.append(f"                else if (wr_en_{reg.reg_name}[i] == 1'b1)")
                        code.append(f"                    {field.name}[i] <= (~wr_data[{field.bits}] | ~msk[{field.bits}]) & {field.name}[i];")
                    code.append("            end")
                    code.append("        end")
                    code.append("    end")
                    code.append("endgenerate")
            # Read data logic
            code.append( "//============================================================================")
            code.append(f"// rd_data_nxt_{reg.reg_name}")
            code.append("//============================================================================")
            if not ginter_j_declared:
                code.append("integer j;")
                ginter_j_declared = True
            code.append(f"always @(*) begin")
            code.append(f"    rd_data_nxt_{reg.reg_name}[31:0]  = 32'h0;")
            code.append(f"    for(j = 0; j <= {reg.var_val}; j = j + 1) begin:rdata_loop_{reg.reg_name}")
            code.append(f"        if (reg_addr[{_addr_width-1}:0] == {_addr_width}'h{reg.offset} + {_addr_width}'h{reg.var_step} * j) begin")
            for field in reg.fields:
                code.append(f"            rd_data_nxt_{reg.reg_name}[{field.bits}] = {field.name}[j][{field.bits_size -1}:0];")
            code.append("        end")
            code.append("    end")
            code.append("end\n")    
        else:# Non-variable registers                  
            for field in reg.fields:
                if field.sw_access != "RO": 
                    code.append( "//============================================================================")
                    code.append(f"// {field.name} addr:{_addr_width}'h{reg.offset} type:{field.sw_access} bits:[{field.bits}] default:{field.default}")
                    code.append("//============================================================================")
                    code.append("always @(posedge clk or negedge rst_n) begin")
                    code.append("    if (!rst_n)")
                    if field.bits_size == 1:
                        code.append(f"        {field.name} <= {field.default};")
                    else:
                        code.append(f"        {field.name}[{field.bits_size-1}:0] <= {field.default};")
                    code.append("    else begin")
                    if field.sw_access == "RW":
                        code.append(f"        if (wr_en_{reg.reg_name} == 1'b1)")
                        if field.bits_size == 1:
                            code.append(f"            {field.name} <= ({field.name} & ~msk[{field.bits}]) | (wr_data[{field.bits}] & msk[{field.bits}]);")
                        else:
                            code.append(f"            {field.name}[{field.bits_size-1}:0] <= ({field.name}[{field.bits_size-1}:0] & ~msk[{field.bits}]) | (wr_data[{field.bits}] & msk[{field.bits}]);")
                    elif field.sw_access == "W1P":
                        code.append(f"        if (wr_en_{reg.reg_name} == 1'b1)")
                        code.append(f"            {field.name} <= wr_data[{field.bits}] & msk[{field.bits}];")
                        code.append("        else")
                        code.append(f"            {field.name} <= 1'b0;")
                    elif field.sw_access == "W1C":
                        code.append(f"        if ({field.name}_hw_en == 1'b1)")
                        code.append(f"            {field.name} <= {field.name}_hw_val;")
                        code.append(f"        else if (wr_en_{reg.reg_name} == 1'b1)")
                        code.append(f"            {field.name} <= (~wr_data[{field.bits}] | ~msk[{field.bits}]) & {field.name};")
                    
                    code.append("    end")
                    code.append("end\n")
    
    # Read data logic
    code.append("//============================================================================")
    code.append("// next read data")
    code.append("//============================================================================")
    code.append("always @(*) begin")
    code.append("    rd_data_nxt[31:0] = 32'h0;")
    code.append(f"    case(reg_addr[{_addr_width-1}:0])")
    
    for reg in registers:
        if reg.var is None:
            addr = reg.offset
            code.append(f"    {_addr_width}'h{addr}: begin")
            for field in reg.fields:
                if field.sw_access == "W1P":
                    code.append(f"        rd_data_nxt[{field.bits}] = {field.bits_size}'h0;")
                else:
                    if field.bits_size == 1:
                        code.append(f"        rd_data_nxt[{field.bits}] = {field.name};")
                    else:   
                        code.append(f"        rd_data_nxt[{field.bits}] = {field.name}[{field.bits_size-1}:0];")
            code.append("    end")
    all_var_none = all(reg.var is None for reg in registers)
    if all_var_none:
        code.append("    default:")
        code.append("        rd_data_nxt = 32'h0;")
    else:
        # Default read data for variable registers
        default_signal = []  
        for reg in registers:
            if reg.var is not None:
                default_signal.append(f"rd_data_nxt_{reg.reg_name} | ")
        # Remove last " | "
        default_signal = "".join(default_signal)[:-3]
        code.append( "    default: ")
        code.append(f"        rd_data_nxt = {default_signal};")
    code.append("    endcase")
    code.append("end\n")
    
    # Final read data register
    code.append("//============================================================================")
    code.append("// read data")
    code.append("//============================================================================")
    code.append("always @(posedge clk or negedge rst_n) begin")
    code.append("    if (!rst_n)")
    code.append("        rd_data[31:0] <= 32'h0; ")
    code.append("    else if(rd_en)")
    code.append("        rd_data[31:0] <= rd_data_nxt[31:0];")
    code.append("end\n")
    
    code.append("endmodule")
    
    return "\n".join(code)

# if __name__ == "__main__":
#     module_info, registers = parse_excel("sys_reg.xlsx",parallel=False)
#     verilog_code = generate_verilog(module_info, registers, "sys_reg")
#     with open("sys_pcie_reg.v", "w") as f:
#         f.write(verilog_code)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Generate Verilog registers from Excel')
    parser.add_argument('input', help='Input Excel file')
    parser.add_argument('-o', '--output', help='Output Verilog file name. Default by module name')
    parser.add_argument('-p', '--parallel', action='store_true', help='Enable parrllel implement for loop. Default by array implement')
    args = parser.parse_args()

    module_info, registers = parse_excel(args.input, parallel=args.parallel)
    output_file = args.output if args.output else f"{module_info['module']}.v"
    
    verilog_code = generate_verilog(module_info, registers, output_file)
    with open(output_file, 'w') as f:
        f.write(verilog_code)
    print(f"Successfully generated {output_file}")
